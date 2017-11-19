from random import randint
from openpyxl import load_workbook

wb = load_workbook('Options_Outcomes.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

def menuMaker(region, area, percentile, silence=False):
    everythingRow = []
    levelRow = []
    regionRow = []
    results = {}
    if area == None:
        area = 'INGENIERIA ARQUITECTURA URBANISMO ECONOMIA ADMINISTRACION CONTADURIA AFINES' #All Areas in case one is not input
    for i in range(2, len(ws['a'])):
        if (ws.cell(row=i, column=3).value) == region:  #RegionCheck
            if area in (ws.cell(row=i, column=5).value):    #AreaCheck
                try:    #IMPORTANT NOTE: This code disregards the percentiles; many of them are missing in Options_Outcomes
                    if percentile >= (ws.cell(row=i, column=2).value): 
                        everythingRow.append(i), levelRow.append(i)
                except TypeError: everythingRow.append(i), levelRow.append(i)

    '''
    MODDED VERION OF PARSING ABOVE(NOT FINISHED)
    
    for i in range(2, len(ws['a'])):
        if (ws.cell(row=i, column=3).value) == region:
            regionRow.append(i)
            if percentile >= (ws.cell(row=i, column=2).value):
                if area in (ws.cell(row=i, column=5).value):
                    everythingRow.append(i)
                
    ''' 
    
    temp = []
    for item in levelRow:
        temp.append(ws.cell(row=item, column=4).value)
     
    print("Uni:", temp.count('Universitaria'))
    print("Tec:", temp.count('Tecnológica'))
    print("FTP:", temp.count('Formación Técnica Profesional'))
    '''
    # This is a Work in Progress of the Random Object Generator, in hopes that it follows 
    # the needed criteria of:
    #    6 IN TOTAL:
    #       2 OF EACH LEVEL:
    #          1 OF SAME REGION
    #          1 OF ANY REGION
    # These have to be somewhat random choices too.
    
    if temp.count('Formación Técnica Profesional') and temp.count('Universitaria') and temp.count('Tecnológica') >= 2:
        s, t, a = [], len(everythingRow), 0
        r = 6
        while a != r:
            p = randint(0, t) - 2
            try:
                s.append(everythingRow.pop(p))
                a += 1
            except: a = a
    else:
        s, t, a = [], len(everythingRow), 0
        r = len(everythingRow) if len(everythingRow) < 6 else 6
        while a != r:
            p = randint(0, t) - 2
            try:
                s.append(everythingRow.pop(p))
                a += 1
            except: a = a
    '''
    #Random Object Generator
    s, t, a = [], len(everythingRow), 0
    r = len(everythingRow) if len(everythingRow) < 6 else 6
    while a != r:
        p = randint(0, t) - 2
        try:
            s.append(everythingRow.pop(p))
            a += 1
        except: a = a
        
    #Result Generator
    for e in s:
        results[r] = (ws.cell(row=e, column=6).value), (ws.cell(row=e, column=4).value), (ws.cell(row=e, column=7).value), str((ws.cell(row=e, column=8).value))
        r -= 1
    if silence != True:
        for q in range(1, a + 1):
            try:
                print("""
Institution: {} 
Level: {}
Field: {} 
Pay: {} pesos.""".format(results[q][0], results[q][1], results[q][2],results[q][3]))
                print()
            except KeyError: raise KeyError
        return
    else: return results
    print(levelRow)


# EXAMPLE OF USAGE: 
#menuMaker('MAGDALENA', 'BELLAS ARTES', 96)
    
menuMaker('MAGDALENA', 'BELLAS ARTES', 96)