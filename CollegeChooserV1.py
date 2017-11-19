from random import randint
from openpyxl import load_workbook

wb = load_workbook('Options_Outcomes.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

def menuMaker(region, area, percentile, silence=False):
    collegeRow = []
    results = {}
    if area == None:
        area = 'INGENIERIA ARQUITECTURA URBANISMO ECONOMIA ADMINISTRACION CONTADURIA AFINES' #All Areas in case one is not input
    for i in range(2, len(ws['a'])):
        if (ws.cell(row=i, column=3).value) == region:  #RegionCheck
            if area in (ws.cell(row=i, column=5).value):    #AreaCheck
                try:    #IMPORTANT NOTE: This code disregards the percentiles; many of them are missing in Options_Outcomes
                    if percentile >= (ws.cell(row=i, column=2).value): collegeRow.append(i)
                except TypeError: collegeRow.append(i)
    s, t, a = [], len(collegeRow), 0
    r = len(collegeRow) if len(collegeRow) < 6 else 6
    while a != r:
        p = randint(0, t) - 2
        try:
            s.append(collegeRow.pop(p))
            a += 1
        except: a = a
    for e in s:
        results[r] = (ws.cell(row=e, column=6).value), (ws.cell(row=e, column=7).value), str((ws.cell(row=e, column=8).value))
        r -= 1
    if silence != True:
        for q in range(1, a + 1):
            try:
                print("""
Institution: {} 
Field: {} 
Pay: {} pesos.""".format(results[q][0],results[q][1],results[q][2]))
                print()
            except KeyError: raise KeyError
        return
    else: return results



# EXAMPLE OF USAGE: 
#('CALDAS', 'BELLAS ARTES', 14)
menuMaker('MAGDALENA', 'BELLAS ARTES', 96)
