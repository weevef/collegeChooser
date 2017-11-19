from random import randint
from openpyxl import load_workbook

wb = load_workbook('Options_Outcomes.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

def menuMaker(region, area, percentile, silence=False):
    
    allRow, regionRow, percentileRow, areaRow = [], [], [], []
    collegeRow = []
    levels = {1:[], 2:[], 3:[]} # Tecnológica = 1, Universitaria = 2, Formación Técnica Profesional = 3
    results = {}
    if area == None:
        area = 'INGENIERIA ARQUITECTURA URBANISMO ECONOMIA ADMINISTRACION CONTADURIA AFINES' #All Areas in case one is not input
    for i in range(2, len(ws['a'])):
        if (ws.cell(row=i, column=3).value) == region:  #RegionCheck
            if area in (ws.cell(row=i, column=5).value):    #AreaCheck
                try:    #IMPORTANT NOTE: This code disregards the percentiles; many of them are missing in Options_Outcomes
                    if percentile >= (ws.cell(row=i, column=2).value): collegeRow.append(i)
                except TypeError: collegeRow.append(i)
    
    for i in range(2, len(ws['a'])):
        allRow.append(i)
        if ws.cell(row=i, column=3).value == region: regionRow.append(i)
        try:
            if ws.cell(row=i, column=2).value <= percentile: percentileRow.append(i)
        except TypeError: percentileRow.append(i)
        if ws.cell(row=i, column=5).value == area: areaRow.append(i)
        if ws.cell(row=i, column=4).value == "Tecnológica": levels[1].append(i)
        elif ws.cell(row=i, column=4).value == "Universitaria": levels[2].append(i)
        elif ws.cell(row=i, column=4).value == "Formación Técnica Profesional": levels[3].append(i)
    
    test = []
    for u in allRow:
        if u in regionRow and u in percentileRow and u in areaRow:
            test.append(u)
            
    if len(regionRow) >= 2 and len(percentileRow) >= 2 and len(areaRow) >=2:
        print(allRow in test)

    #print(regionRow, '\n Percentile: \n', percentileRow, '\n Area \n',areaRow)



    '''
    for a in allRow:
        while 0 in full:
            break
    '''
    return results

menuMaker('MAGDALENA', 'BELLAS ARTES', 96)