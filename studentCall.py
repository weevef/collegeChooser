from CollegeChooserV2 import menuMaker
from openpyxl import load_workbook

wb = load_workbook('Options_Outcomes.xlsx')
mu = wb.get_sheet_by_name('MockUp')

sl = {}
for i in range(2, len(mu['a'])):
    x = []
    x.append(mu.cell(row=i, column=3).value)
    x.append(mu.cell(row=i, column=4).value)
    x.append(mu.cell(row=i, column=2).value)
    sl[i] = x

howmany = 6
for y in range(2, howmany):
    print("KID ID:", y)
    menuMaker(sl[y][0], sl[y][1], sl[y][2], silence=True)
    print()
